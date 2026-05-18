from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "web" / "data" / "app-data.js"
BASE_FILE = ROOT / "base_jtudela_2026-05-11_1510.xlsx"
CARTOLA_FILE = ROOT / "cartola_abril.xlsx"


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def money(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    digits = re.sub(r"[^\d-]", "", str(value))
    return int(digits) if digits not in ("", "-") else 0


def as_text(value: Any) -> str:
    return str(clean(value)).strip()


def load_base() -> list[dict[str, Any]]:
    wb = load_workbook(BASE_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [as_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = {headers[i]: clean(row[i]) if i < len(row) else "" for i in range(len(headers))}
        correos = [as_text(raw.get(f"correo_{i}")) for i in range(1, 7)]
        telefonos = [as_text(raw.get(f"telefono_{i}")) for i in range(1, 21)]
        correos = [c for c in correos if c and c.lower() != "none"]
        telefonos = [t for t in telefonos if t and t.lower() != "none"]

        saldo_capital = money(raw.get("saldo_capital"))
        deuda_total = money(raw.get("deuda_total"))
        oferta = round(saldo_capital * 0.5)

        rows.append(
            {
                "id": as_text(raw.get("id_rem")),
                "rutDeudor": as_text(raw.get("rut_deudor")),
                "rutTitular": as_text(raw.get("rut_titular")),
                "nombreTitular": as_text(raw.get("nombre_titular")),
                "rutAlumno": as_text(raw.get("rut_alumno")),
                "nombreAlumno": as_text(raw.get("nombre_alumno")),
                "usuario": as_text(raw.get("usuario")),
                "asignacion": as_text(raw.get("asignacion")),
                "equipo": as_text(raw.get("equipo")),
                "estado": as_text(raw.get("estado")) or "Pendiente",
                "cartera": as_text(raw.get("cartera")),
                "tramo": as_text(raw.get("tramo_deuda")),
                "fechaEmision": as_text(raw.get("fecha_emision")),
                "saldoCapital": saldo_capital,
                "deudaTotal": deuda_total,
                "interes": money(raw.get("interes")),
                "gastoCobranza": money(raw.get("gasto_cobranza")),
                "montoOferta": oferta,
                "atrasoGestion": as_text(raw.get("atraso_gestion")),
                "ultimaGestion": as_text(raw.get("ultima_gestion")),
                "tipoContacto": as_text(raw.get("tipo_contacto")),
                "resultado": as_text(raw.get("resultado")),
                "observacion": as_text(raw.get("observacion")),
                "proximaGestion": as_text(raw.get("proxima_gestion")),
                "ubicabilidad": as_text(raw.get("ubicabilidad")),
                "direccion": as_text(raw.get("direccion")),
                "comuna": as_text(raw.get("comuna")),
                "region": as_text(raw.get("region")),
                "rol": as_text(raw.get("rol")),
                "tribunal": as_text(raw.get("tribunal")),
                "correos": correos[:6],
                "telefonos": telefonos[:8],
                "telValidado": as_text(raw.get("tel_validado")),
                "scoreContactabilidad": min(100, len(correos) * 15 + len(telefonos) * 10 + (20 if as_text(raw.get("tel_validado")).lower() in ("si", "sí", "validado") else 0)),
            }
        )
    return rows


def load_cartola() -> list[dict[str, Any]]:
    wb = load_workbook(CARTOLA_FILE, read_only=True, data_only=True)
    movements: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if ws.title.lower() == "sumas":
            continue
        headers = [as_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = {headers[i]: clean(row[i]) if i < len(row) else "" for i in range(len(headers))}
            amount = money(raw.get("Ingreso (+)") or raw.get("Depósitos / Abonos"))
            if amount <= 0:
                continue
            fecha = raw.get("Fecha de transacción") or raw.get("Fecha") or raw.get("Contable")
            movements.append(
                {
                    "fuente": ws.title,
                    "fecha": as_text(fecha),
                    "hora": as_text(raw.get("Hora transacción")),
                    "monto": amount,
                    "nombre": as_text(raw.get("Nombre")),
                    "rut": as_text(raw.get("RUT")),
                    "codigo": as_text(raw.get("Código Transferencia") or raw.get("N° Operación") or raw.get("Código de transacción")),
                    "glosa": as_text(raw.get("Glosa detalle") or raw.get("Descripción") or raw.get("Comentario transferencia")),
                    "tipo": as_text(raw.get("Tipo de transacción") or raw.get("Tipo")),
                }
            )
    return movements


def build_summary(records: list[dict[str, Any]], movements: list[dict[str, Any]]) -> dict[str, Any]:
    states = Counter(r["estado"] for r in records)
    executives = Counter(r["usuario"] or "Sin usuario" for r in records)
    regions = Counter(r["region"] or "Sin región" for r in records)
    contact = {
        "conCorreo": sum(1 for r in records if r["correos"]),
        "conTelefono": sum(1 for r in records if r["telefonos"]),
        "sinDatos": sum(1 for r in records if not r["correos"] and not r["telefonos"]),
    }
    movement_by_source: dict[str, int] = defaultdict(int)
    for m in movements:
        movement_by_source[m["fuente"]] += m["monto"]

    return {
        "totalRegistros": len(records),
        "saldoCapital": sum(r["saldoCapital"] for r in records),
        "deudaTotal": sum(r["deudaTotal"] for r in records),
        "montoOferta": sum(r["montoOferta"] for r in records),
        "comision25SobreOferta": round(sum(r["montoOferta"] for r in records) * 0.25),
        "estados": states.most_common(),
        "ejecutivos": executives.most_common(),
        "regiones": regions.most_common(8),
        "contactabilidad": contact,
        "cartola": {
            "movimientos": len(movements),
            "montoIngresos": sum(m["monto"] for m in movements),
            "porFuente": sorted(movement_by_source.items()),
        },
    }


def main() -> None:
    records = load_base()
    movements = load_cartola()
    data = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "businessRules": {
            "discountRate": 0.5,
            "offerFormula": "saldo_capital * 50%",
            "commissionRate": 0.25,
            "bankAccount": "Banco BCI - Comercial Remesa SpA - RUT 76.976.117-9 - Cuenta Corriente 27826341",
        },
        "summary": build_summary(records, movements),
        "debtors": records,
        "bankMovements": movements[:500],
    }
    OUT.write_text("window.ABG_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(f"Generated {OUT} with {len(records)} debtors and {len(movements)} bank movements")


if __name__ == "__main__":
    main()
