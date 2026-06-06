import hashlib
import io
import os
import re
import secrets
from datetime import datetime, timezone

import openpyxl

from import_pending import dedupe_by, map_contacts, map_debtor, normalize_header, supabase_upsert
from setup_drive import drive_service


BASE_SHEET_ID = os.getenv("GOOGLE_SHEETS_AIEP_BASE_ID") or "1JLprSdfbtg2MdPZbjQklsuvWcb4Vz696uTll0laGnFw"


def normalize_username(value):
    import unicodedata

    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", ".", text.lower()).strip(".")
    return text[:42] or "ejecutivo"


def hash_password(password, salt):
    return hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()


def download_workbook():
    service = drive_service()
    request = service.files().export_media(
        fileId=BASE_SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return request.execute()


def sheet_rows(workbook, sheet_name):
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header or "").strip() for header in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
    ]


def value(row, name):
    target = normalize_header(name)
    for key, cell in row.items():
        if normalize_header(key) == target:
            return cell
    return ""


def active_value(raw):
    return not re.match(r"^(no|false|inactivo|0)$", str(raw or "").strip(), re.I)


def sync_base(rows):
    mapped = [(row, map_debtor(row)) for row in rows]
    debtors = dedupe_by(
        [debtor for _, debtor in mapped if debtor["rut_titular_normalizado"] or debtor["rut_alumno_normalizado"]],
        lambda item: item["id"],
    )
    contacts = dedupe_by(
        [contact for row, debtor in mapped for contact in map_contacts(row, debtor["id"])],
        lambda item: f'{item["debtor_id"]}|{item["type"]}|{item["value"].lower()}',
    )
    supabase_upsert("debtors", debtors, "id")
    supabase_upsert("contacts", contacts, "debtor_id,type,value")
    return len(debtors), len(contacts)


def sync_assigned(rows):
    users = []
    now = datetime.now(timezone.utc).isoformat()
    for row in rows:
        name = str(value(row, "Nombre") or "").strip()
        if not name or name.lower().startswith("total"):
            continue
        username = normalize_username(value(row, "Usuario") or name)
        password = str(value(row, "Contrasena") or value(row, "Contraseña") or "123456")
        role = str(value(row, "Rol") or "callcenter").strip().lower()
        if role not in {"callcenter", "jefatura", "informatico"}:
            role = "callcenter"
        salt = secrets.token_hex(16)
        users.append({
            "username": username,
            "display_name": name,
            "role": role,
            "assignment_name": name if role == "callcenter" else "",
            "active": active_value(value(row, "Activo") or "SI"),
            "password_hash": hash_password(password, salt),
            "password_salt": salt,
            "password_changed_at": now,
            "updated_at": now,
            "created_at": now,
        })
    supabase_upsert("app_users", users, "username")
    return len(users)


def main():
    for name in ["SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY"]:
        if not os.getenv(name):
            raise RuntimeError(f"Falta {name}")

    workbook = openpyxl.load_workbook(io.BytesIO(download_workbook()), data_only=True, read_only=True)
    if "Base" not in workbook.sheetnames or "Asignados" not in workbook.sheetnames:
        raise RuntimeError("AIEP_BASE_TOTAL debe tener hojas Base y Asignados")

    debtors, contacts = sync_base(sheet_rows(workbook, "Base"))
    users = sync_assigned(sheet_rows(workbook, "Asignados"))
    print(f"OK AIEP_BASE_TOTAL importado: {debtors} deudores, {contacts} contactos, {users} perfiles")


if __name__ == "__main__":
    main()
