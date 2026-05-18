import csv
import io
import os
import re
from datetime import datetime, timezone

import openpyxl
import requests

from setup_drive import ROOT_FOLDER_ID, drive_service, ensure_path


SUPPORTED_MIME_TYPES = {
    "application/vnd.google-apps.spreadsheet",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "text/plain",
}


def normalize_header(value):
    import unicodedata

    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", "_", text.strip().lower())


def normalize_rut(value):
    return re.sub(r"[^0-9Kk]", "", str(value or "")).upper()


def text(value):
    return str(value or "").strip()


def money(value):
    if isinstance(value, (int, float)):
        return round(value)
    raw = str(value or "").strip().replace(" ", "")
    if not raw:
        return 0
    if re.search(r",\d{1,2}$", raw):
        raw = re.sub(r",\d{1,2}$", "", raw)
    elif re.search(r"\.\d{1,2}$", raw) and not re.search(r"\.\d{3}(\D|$)", raw):
        raw = re.sub(r"\.\d{1,2}$", "", raw)
    cleaned = re.sub(r"[^0-9-]", "", raw)
    return int(cleaned or "0")


def find_value(row, aliases):
    normalized = {normalize_header(alias) for alias in aliases}
    for key, value in row.items():
        if normalize_header(key) in normalized:
            return value
    return ""


def split_values(value):
    return [
        item.strip()
        for item in re.split(r"[;,|/\n\r]+", str(value or ""))
        if item.strip() and item.strip().lower() != "none"
    ]


def numbered_values(row, prefix, limit):
    values = []
    for index in range(1, limit + 1):
        value = text(find_value(row, [f"{prefix}_{index}"]))
        if value and value.lower() != "none":
            values.append(value)
    return values


def map_debtor(row):
    rut_titular = text(find_value(row, ["rut titular", "rut_titular", "rut deudor", "rut_deudor"]))
    rut_alumno = text(find_value(row, ["rut alumno", "rut_alumno", "rut estudiante", "rut_estudiante"]))
    debtor_id = text(find_value(row, ["id", "id rem", "id_rem", "codigo", "operacion", "remesa"]))
    if not debtor_id:
        debtor_id = f"AIEP {normalize_rut(rut_titular)} {normalize_rut(rut_alumno)}"

    saldo_capital = money(find_value(row, ["saldo capital", "capital", "saldo_capital", "monto deuda", "monto_deuda"]))
    intereses_mora = money(find_value(row, ["intereses mora", "interes", "interes mora", "intereses_mora"]))
    gastos_cobranza = money(find_value(row, ["gastos cobranza", "gasto cobranza", "gasto_cobranza", "gastos_cobranza"]))
    deuda_total = money(find_value(row, ["deuda total", "total", "deuda_total"])) or saldo_capital + intereses_mora + gastos_cobranza

    return {
        "id": debtor_id.strip(),
        "rut_titular": rut_titular,
        "rut_titular_normalizado": normalize_rut(rut_titular),
        "nombre_titular": text(find_value(row, ["nombre titular", "titular", "nombre_titular", "nombre deudor", "nombre_deudor", "deudor"])),
        "rut_alumno": rut_alumno,
        "rut_alumno_normalizado": normalize_rut(rut_alumno),
        "nombre_alumno": text(find_value(row, ["nombre alumno", "alumno", "nombre_alumno", "nombre estudiante", "nombre_estudiante", "estudiante"])),
        "estado": text(find_value(row, ["estado"])) or "ACUERDO ROTO",
        "cartera": text(find_value(row, ["cartera"])),
        "tramo": text(find_value(row, ["tramo", "tramo_deuda"])),
        "region": text(find_value(row, ["region"])),
        "comuna": text(find_value(row, ["comuna"])),
        "direccion": text(find_value(row, ["direccion"])),
        "rol": text(find_value(row, ["rol"])),
        "tribunal": text(find_value(row, ["tribunal"])),
        "saldo_capital": saldo_capital,
        "intereses_mora": intereses_mora,
        "gastos_cobranza": gastos_cobranza,
        "deuda_total": deuda_total,
        "monto_oferta": money(find_value(row, ["monto oferta", "oferta", "monto_oferta"])),
    }


def map_contacts(row, debtor_id):
    emails = (
        split_values(find_value(row, ["correo", "correos", "mail", "email", "e-mail"]))
        + numbered_values(row, "correo", 10)
        + numbered_values(row, "email", 10)
        + numbered_values(row, "mail", 10)
    )
    phones = (
        split_values(find_value(row, ["telefono", "telefonos", "celular", "celulares", "fono"]))
        + numbered_values(row, "telefono", 20)
        + numbered_values(row, "celular", 20)
        + numbered_values(row, "fono", 20)
    )

    contacts = []
    for value in emails:
        contacts.append({"debtor_id": debtor_id, "type": "correo", "value": value, "status": "sin_validar", "note": ""})
    for value in phones:
        cleaned = re.sub(r"\D", "", value)
        if cleaned:
            contacts.append({"debtor_id": debtor_id, "type": "telefono", "value": cleaned, "status": "sin_validar", "note": ""})
    return contacts


def dedupe_by(rows, key_fn):
    seen = set()
    output = []
    for row in rows:
        key = key_fn(row)
        if not key or key in seen:
            continue
        seen.add(key)
        output.append(row)
    return output


def download_file(service, file):
    if file.get("mimeType") == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=file["id"],
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=file["id"], supportsAllDrives=True)
    return request.execute()


def parse_rows(file_name, content):
    name = file_name.lower()
    if name.endswith(".csv"):
        decoded = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(decoded)))

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [text(header) for header in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
    ]


def supabase_headers():
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json; charset=utf-8",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }


def supabase_upsert(table, rows, conflict, batch_size=100):
    if not rows:
        return 0
    base_url = os.environ["SUPABASE_URL"].rstrip("/")
    url = f"{base_url}/rest/v1/{table}?on_conflict={conflict}"
    sent = 0
    for index in range(0, len(rows), batch_size):
        batch = rows[index:index + batch_size]
        response = requests.post(url, headers=supabase_headers(), json=batch, timeout=120)
        if response.status_code >= 400:
            raise RuntimeError(f"Supabase {table}: {response.status_code} {response.text}")
        sent += len(batch)
    return sent


def log_import(file, folder_path, status, rows_read=0, debtors=0, contacts=0, error_message=None):
    payload = {
        "drive_file_id": file["id"],
        "drive_file_name": file["name"],
        "drive_folder_path": folder_path,
        "status": status,
        "rows_read": rows_read,
        "debtors_upserted": debtors,
        "contacts_upserted": contacts,
        "error_message": error_message,
        "processed_at": datetime.now(timezone.utc).isoformat(),
    }
    supabase_upsert("drive_imports", [payload], "drive_file_id", 1)


def move_file(service, file_id, target_folder_id):
    file = service.files().get(fileId=file_id, fields="parents", supportsAllDrives=True).execute()
    previous_parents = ",".join(file.get("parents", []))
    service.files().update(
        fileId=file_id,
        addParents=target_folder_id,
        removeParents=previous_parents,
        fields="id,parents",
        supportsAllDrives=True,
    ).execute()


def main():
    for name in ["SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY"]:
        if not os.getenv(name):
            raise RuntimeError(f"Falta {name}")

    service = drive_service()
    pending_id = ensure_path(service, ROOT_FOLDER_ID, "01_Fuentes_Cartera/Pendientes")
    processed_path = f"01_Fuentes_Cartera/Procesados/{datetime.now(timezone.utc).strftime('%Y-%m')}"
    processed_id = ensure_path(service, ROOT_FOLDER_ID, processed_path)
    rejected_id = ensure_path(service, ROOT_FOLDER_ID, "01_Fuentes_Cartera/Rechazados")

    pending = service.files().list(
        q=f"'{pending_id}' in parents and trashed=false",
        fields="files(id,name,mimeType)",
        pageSize=50,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])

    failed = 0
    for file in pending:
        if file.get("mimeType") == "application/vnd.google-apps.folder":
            continue
        try:
            if file.get("mimeType") not in SUPPORTED_MIME_TYPES and not file["name"].lower().endswith((".xlsx", ".xls", ".csv")):
                raise RuntimeError(f"Tipo no soportado: {file.get('mimeType')}")
            log_import(file, "01_Fuentes_Cartera/Pendientes", "procesando")
            rows = parse_rows(file["name"], download_file(service, file))
            mapped = [(row, map_debtor(row)) for row in rows]
            debtors = dedupe_by(
                [debtor for _, debtor in mapped if debtor["rut_titular_normalizado"] or debtor["rut_alumno_normalizado"]],
                lambda item: item["id"],
            )
            contacts = dedupe_by(
                [contact for row, debtor in mapped for contact in map_contacts(row, debtor["id"])],
                lambda item: f'{item["debtor_id"]}|{item["type"]}|{item["value"].lower()}',
            )
            supabase_upsert("debtors", debtors, "id")
            supabase_upsert("contacts", contacts, "debtor_id,type,value")
            log_import(file, processed_path, "procesado", len(rows), len(debtors), len(contacts))
            move_file(service, file["id"], processed_id)
            print(f'Importado {file["name"]}: {len(debtors)} deudores, {len(contacts)} contactos')
        except Exception as exc:
            failed += 1
            message = str(exc)
            log_import(file, "01_Fuentes_Cartera/Rechazados", "rechazado", error_message=message)
            move_file(service, file["id"], rejected_id)
            print(f'Rechazado {file["name"]}: {message}')

    if failed:
        raise SystemExit(f"{failed} archivo(s) rechazado(s)")


if __name__ == "__main__":
    main()
